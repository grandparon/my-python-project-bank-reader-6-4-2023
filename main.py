import random

import xlsxwriter

#import pandas
#Revisions 1.0 Todat 2/11/: playing hith github
# revision 2.0 3/21 Added SBF (Standard Banking Format) and removed redundent sections of main for each bank


#*****************************************************************************
# **************************** GLOBAL VARIABLE DECLARATIONS  *****************
#*****************************************************************************
magic_number_quit_and_dont_save= int(20)
magic_number_quit_and_save= int(21)
bank_data = {}
row_count =rows_this_session = 0
new_rows_completed = 0
sums_info_dict = {"date":"str_date"}
pass
save_file = "c:\python-write-data\saved_bank_data.xlsx"
saved_graph_file = "c:\python-write-data\saved_graph.xlsx"

statement_directory = "f:Libraries-System-Win10/Downloads"
sbf_dict = {}
gws = 0

statement_headings_dict = {"STORE": ["NV",3,7,2], "AMOUNT": ["NV",5,4,4], "DATE": ["NV",1,2,0], "CATEGORY": ["NV",4, 8, 1],"PAYMENT": ["NV",6, 6, 6]}
bank_file_name_dict = {"CAPITAL_ONE":"transaction_download." ,"US_BANK":"Checking" ,"FIRST_TECH": "ExportedTransactions"}
sbf_headings_dict = {"BANK NAME": 0, "DATE": 1, "AMOUNT": 2, "CATEGORY":4}
month_wanted = 1
temp_cat_sums_from_file = {}
graphs_sheet_name = "Graphs"
cat_sheet_name = "category sums"
# *****************************************************************************


import os
import openpyxl
import csv
import json
import pandas as pd
import os.path
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BubbleChart,  BarChart, PieChart, ProjectedPieChart, Series, Reference
from openpyxl.chart.series import DataPoint
import numbers
from enum import Enum
import datetime as dt
from datetime import date
import calendar
import sys
import re
import matplotlib.pyplot as plt









us_bank_regex = "nop"

bank_dictionary = {}
bank_name_dict = { 0: "EXIT", 1: "CAPITAL_ONE", 2: "FIRST_TECH", 3: "US_BANK"}
#category_dict = {1:["Boat",0], 2: ["Food",0], 3:["Dining",0], 4: ["House",0], 5:["Travel",0], 6:["Utilities",0], 7:["Auto",0], 8:["Health",0],
#                 9:["Deductable",0], 10: ["Recreation",0], 11: ["Unknown",0], 12: ["Pay Credit Card",0], 13: ["Deposits to USB",0],
#                 14:["Money Transfers",0], 15: ["Interest",0], 16:["Trivials",0],17: ["Taxes",0], 18: ["Date", "STR"], 19: ["Bank", "STR"], 20: ["Print bank_data",0], 21: ["Print Sums",0],
#                 22: ["Quit but don't Save",0], 23: ["Save and Quit",0]}


category_dict_2 = {"[A]uto": [0],
                   "[B]oat": [0],
                   "[C]redit Card Payment": [0],
                   "[D]eposits to US Bank": [0],
                   "[E]verything":[0],
                   "[F]ood and resturents": [0],
                    "[G]iving":[0],
                   "[H]ealth":[0],
                   "[I]nterest": [0],
                   "[J]unk": [0],
                   "[K]not Valid":[0],
                   "[L]ame Duck": [0],
                    "[M]oney transfers":[0],
                   "[N]ot This One" : [0],
                    "[O]ur House": [0],
                    "[P]oor Farm":[0],
                    "[Q]uit and Save":[-1],
                    "[R]ecreation":[0],
                    "[S]mall stuff":[0],
                    "[T]ravel": [0],
                    "[U]tilities":[0],
                   "[V]Not a [V]alid Entry":[0],
                   "[W]Not a [W]alid Entry ": [0],
                   "Ta[X]es": [0],
                    "[Y]Unknown": [0],
                   "Quit[Z]-Don't Save": [0],
                   "[]lank1":[0],
                   "[]lank2":[0]
                  }
def month_intro():
    print ("statement for what month?")
    for i in range(1, 12):
        print(f"{i}: {calendar.month_name[i]}")
    mw = input ("\n")
    if mw.isdigit():
        if 1 <= int(mw) <= 12:
            return (int(mw))
    else:
        return ("NOT A VALID MONTH")
def intro():
    selection = "NONE"
    print("1 for Capital One")
    print("2 for First Tech")
    print ("3 for US Bank")
    print ("4 for a sum of costs")
#    print ("5 for a graph")
    print ("0 for exit")
    selection = input("select bank\n")
    selection = int(selection)
    if selection > 0 and selection < 4 :
        return selection, bank_name_dict[int(selection)]
#    elif  selection == 5:
#        g = Graph(save_file)
#        g.new_graph()
    else:
        exit()



def  get_key_by_value(val):
    keys_list = get_list_of_category_keys()
    cat_key = keys_list[val-1]
    return cat_key

def get_list_of_category_keys():
    cat_key_list = list(category_dict_2.keys())
    return (cat_key_list)
class Category_headings(Enum):
    BANK          =   19

class Sbf_enum (Enum):
    COMPANY             = "IS KEY"
    BANK_NAME           = 0
    TRANSACTION_DATE    = 1
    AMOUNT              = 2
    CATEGORY            = 3
    STATEMENT_MONTH     = 4
    STR_CATEGORY        = 5
    PAYMENT             = 6

class First_Tech_format (Enum):
    DATE            = 2
    DEBIT_OR_CREDIT = 3
    AMOUNT          = 4
    COMPANY         = 8
    CATEGORY        = 9


class Bank_Names (Enum):
    NONE                = 0
    CAPITAL_ONE         = 1
    FIRST_TECHNOLOGY    = 2
    US_BANK             = 3

class Num_To_Month (Enum):
    JANUARY         = 1

class SBF:
    def __init__(self):
        pass

class Bank:
    def __init__(self, bank_name:str, bank_num:int, bank_file_name:str, month:float,
    download_directory:str):
        self.bank_name = bank_name
        self.bank_name = bank_num
        self.bank_file_name = bank_file_name
        self.month = month
        self.download_directory = download_directory
    pass

    def csv_to_xl(self):

        for file_name in os.listdir(self.download_directory):
            if file_name.endswith(".csv") and (self.bank_file_name in file_name) == True :
                a_csv_file = statement_directory + '/' + file_name
                wb = openpyxl.Workbook()
                ws = wb.active
                with open(a_csv_file) as f:
                    reader = csv.reader(f, delimiter=",")
                    for row in reader:
                        ws.append(row)
                file_name_with_dot_xlsx = file_name.replace("csv", "xlsx")
                wb.save(statement_directory + "/" + file_name_with_dot_xlsx )
                os.remove(statement_directory + "/" + file_name )

    def get_wb_and_ws (self):
        list_of_bank_files = []
        pass
        for file_name in os.listdir(self.download_directory):
            if file_name.endswith(".xlsx") and (self.bank_file_name in file_name) == True :
                list_of_bank_files.append(file_name)
                wb = openpyxl.load_workbook(self.download_directory + "/" + file_name )
                ws = wb.active
                if bank_name == "FIRST_TECH":
                    cell_date = ws['B3'].value
                    cell_split_date = cell_date.split('/')
                    cell_month = int(cell_split_date[0])
                else:
                    cell_date = ws['A3'].value
                    cell_split_date = cell_date.split('-')
                    cell_month = int(cell_split_date[1])
                if self.month == cell_month:
                    pass
                    return (self.download_directory + "/" + file_name, wb, ws)

    def get_sheet_length(a_bank_file_name, sheet_named):
#test to see sheet exists
        df_dict = pd.read_excel(a_bank_file_name, sheet_name=None)
        if sheet_named in df_dict:
            df1 = pd.read_excel(a_bank_file_name, sheet_name=sheet_named)
            num_rows = len(df1)
            return num_rows
        else:
            return


    def get_statement_length(a_bank_file_name):
            pass
            if os.path.isfile(a_bank_file_name):
                dataframe1 = pd.read_excel(a_bank_file_name)
                pandas_num_rows = dataframe1.shape[0]
                return (ws.min_row,  ws.max_row, ws.min_column,  ws.max_column, pandas_num_rows)
            else:
                print(f" could not find {a_bank_file_name}")
                exit








class Add_categories:

    def input_and_parse_category(self):
        c = -1
        my_set = []
        txt = list(category_dict_2.keys())
        length = len(txt)

        while c == -1 :

            c = input()
            alph_num = c.isalnum()
            if alph_num == False:
                continue

            if c.isdigit():
                mmm = txt[int(c)-1]
                nnn = re.search("\[(\w+)\]", mmm)
                c = nnn.group(1)


            pass
            c_upper = c.upper()
            pattern = "\["+c_upper+"\]"
            i=1
            for item in txt:
                pass
                if re.search(pattern,item):
                #     print (f' match is : {item}')
                #     pass
                #     if item == "[Q]uit and Save":
                #         s = Merge_and_save(save_file, sbf_dict)
                #         s.merge_data()
                #         bank_data.popitem()
                #         s.save()  # line 145
                #         print("Saved")
                #         wb.close()
                #         sys.exit()
                #         pass
                #     if item == "Quit[Z]-Don't Save":
                #         print("==== not saving (from Parse) ==")
                #         sys.exit()
                #         pass
                    return (i)
                i+= 1
            pass





 #           for c in [1,"a", "A"]: return(1)
 #           for c in [2,"b", "B"]: return(2)
 #           for c in [3,"b","B"]: return (3)
    def get_category(self):
        b = Add_categories.input_and_parse_category(self)

        c = b
        txt = len(list(category_dict_2.keys()))
        if 0 <= c < txt :
            return int(c), int(0)


    def print_category_menu(self):
        i=1
#        key_list = get_list_of_category_keys()

        for key, value in category_dict_2.items():
            s_key = key.split("]")[0]
            ds_key = s_key.split("[")[1]
            dds_key = key.split("]")[1]
            colored_key ="["+"\033[1;32m"+ds_key+"\033[1;0m"+"]"+dds_key
            print(f'{i}  {colored_key} {value}')
            i = i + 1
    def parse_cat_nums_and_non_standard_exit_programs(self):
        if 0 < cat_num < 20:
            return ( False, False)
 #       if cat_num == 20:
 #           print ("PRINTING BANK_DATA then continue")
 #           self.print_json()
 #           return (20, True)
 #       elif cat_num == 21:
 #           print ("PRINTING CATEGORY SUMS then continue")
 #           ms = Master_sum(0,0,0)
 #           ms.print_cat_sums()
 #           return (21, True)
        elif cat_num == magic_number_quit_and_dont_save:
            print ("EXITING without saving session")
            exit()
        elif cat_num == magic_number_quit_and_save:
            print ("SAVE and EXIT after this entry")
            return (magic_number_quit_and_save, True)


    def add_new_cat_to_sbf_dict(self, cat_num, new_sbf_dict, cat_num_text_str):
        x = list (new_sbf_dict)[-1]
        pass
        new_sbf_dict[x][Sbf_enum.CATEGORY.value] = cat_num
        new_sbf_dict[x][Sbf_enum.STR_CATEGORY.value] = cat_num_text_str
        pass
    def print_json(self):
        print("=========================")
        print(json.dumps(bank_data, indent=4))
        print("=========================")
class Category_to_xl_column:
    pass
    def read_bank_data(self):
        pass
class Master_sum:
    def __init__(self, new_amount, cat_num, sbf_dict):

        self.new_amount = new_amount
        self.cat_num = cat_num
        self.sbf_dict = sbf_dict
    pass
    def fix_new_amount_signs (self, a_new_sbf_dict_key):
        pass
        amount_to_add = sbf_dict[a_new_sbf_dict_key][sbf_headings_dict["AMOUNT"]]
        if amount_to_add >= 0:
            total_deposits = amount_to_add
            return ( 0, total_deposits,self.cat_num)  #return (deposits, abs(debits) )
        elif amount_to_add < 0:
            amount = round(float(amount_to_add),2)
            pass
            return ( amount, 0, self.cat_num)

    def print_cat_sums (self):
        i=0
        pass
        for i in range (1, 13):
            print (f"sum for {category_dict[i]} is ${category_sums[i]}")

class Summaries():

    def add_money(self):
        sum= 0
        one_key = 0
        global bank_sum
        for keys in bank_data:
            dd  = len (bank_data[keys][0])
            x = bank_data[keys][0]
            for q in range(0,dd):
                one_key = x[q]

            bank_sum = bank_sum + one_key

class Merge_and_save:
    def __init__(self, save_file, sbf_dict, df_g):
        pass
        self.save_file = save_file
        self.sbf_dict = sbf_dict
        self.df_g = df_g

    def copy_chart(self, src_file, dest_file):
        # importing openpyxl module
        import openpyxl as xl;

        # opening the source excel file
        filename = src_file
        wb1 = xl.load_workbook(filename)
        ws1 = wb1.active


        # opening the destination excel file
        filename1 = dest_file
        wb2 = xl.load_workbook(filename1)
        # ws2 = wb2.active
        # ws1 = wb1.worksheets["copied chart"]
        ws2 = wb2.create_sheet("Copied Chart")

        # calculate total number of rows and
        # columns in source excel file
        mr = ws1.max_row
        mc = ws1.max_column

        # copying the cell values from source
        # excel file to destination excel file
        for i in range(1, mr + 1):
            for j in range(1, mc + 1):
                # reading cell value from source excel file
                c = ws1.cell(row=i, column=j)

                # writing the read value to destination excel file
                ws2.cell(row=i, column=j).value = c.value

        # saving the destination excel file
        wb2.save(str(filename1))

    def save_bank_data_to_excel(self):
        pass

    def merge_data(self):
        pass
        bank_data.update(sbf_dict)
        pass
    def save (self, df_g):

        today = date.today()
        td = today.strftime("%d/%m/%y")
        date_dict = {"Bank": [bank_name], "Date": [str(td)]}
        df_date = pd.DataFrame(date_dict)
        with pd.ExcelFile(save_file) as xls:
            if "category sums" in xls.sheet_names:
                pass
                df_12 = pd.read_excel(xls, "category sums")
                pass
            else:
                df_11 = pd.DataFrame(category_dict_2)
                pass
                timanddate = [df_11, df_date]
                df_11 = pd.concat(timanddate, axis=1)
                pass
        pass
        df_from_bank_data = pd.DataFrame.from_dict (bank_data)
        excel_pandas_file = pd.ExcelFile(save_file)
        if "category sums" in excel_pandas_file.sheet_names:
            df_11 = pd.DataFrame(category_dict_2)
            t_and_d = [df_11, df_date]
            df_11 = pd.concat(t_and_d, axis=1)
            frames = [df_12, df_11]
            df_either = pd.concat(frames)
            pass
        else:
            df_either = df_11
            pass

        pass
        # df = pd.DataFrame({"data": [1, 2, 3, 4, 5, 6, 7]})
        # writer = pd.ExcelWriter(save_file, engine="xlsxwriter")
        # df.to_excel(writer)
        #
        # chart = workbook.add_chart({'type': 'column'})
        # chart.add_series({
        #     'values': '=\'Sheet 1\'!$B$2:$B$8',
        #     "name": "My Series's Name"
        # })
        # workshit.insert_chart('D2', chart)
        with pd.ExcelWriter(save_file, engine='xlsxwriter') as writer:

            if "category sums" in excel_pandas_file.sheet_names:
                pass
                df_from_bank_data.T.to_excel(writer)
                df_either.to_excel(writer, sheet_name= 'category sums', startrow=0, startcol=0, header=True)
            else:
                df_from_bank_data.T.to_excel(writer, sheet_name="Sheet1")
                df_11.to_excel(writer, sheet_name="category sums", index=False)


    def add_category(self):
        def __init__(self):
            print ("add a category for")
#===========================================
# Class to for all new statement rows
#===========================================
class New_statement_row:
    def __init__(self, new_row, bank_name, bank_num):
        self.new_row = new_row
        self.bank_name = bank_name
        self.bank_num = bank_num


#----------------------------------------------------------------------------------------------------
#    SBF definition : Bank_Instition   Transaction_date   Statement_month   Store   Amount  Category
# --------------------------


#SBF is:
    def new_sbf_dict(self):
        pur_date = self.new_row[statement_headings_dict["DATE"][bank_num]]
#        purchase_date = pur_date.strftime ('%m/%d/%y')
        pass
        sbf_dict [new_row [statement_headings_dict["STORE"][bank_num]]] = [ bank_name, pur_date,
                                                                            int(round(float(new_row[statement_headings_dict["AMOUNT"][bank_num]]),2)),
                                                                            "RESERVED FOR ASSIGNED CATEGORY NUMBER",
                                                                            month_wanted,
                                                                            new_row[statement_headings_dict["PAYMENT"][bank_num]],
                                                                            "RESERVED", "RESERVED", "RESERVED"]

        return (new_row [statement_headings_dict["STORE"][bank_num]])

    def is_statement_row_in_bd_and_not_in_saved_xl(self, new_statement_key):
        temp_saved_file = {}
        if os.path.isfile(save_file):
            wb1 = load_workbook(save_file)
            ws1 = wb1.active
            pass
            for row in list(ws1.rows)[1:]:
                pass
                temp_saved_file[row[0].value] = [c.value for c in row[1:]]
            temp_saved_file.pop("null", "not_found")
            pass
        else:
            print ("NO SAVED FILE FOUND")
            exit("NO")
            pass
        if new_statement_key in bank_data:
            if new_statement_key in temp_saved_file:
                found_one = True
            else:
                found_one = False

            if (found_one == False):
                print (" found {key} IN bank_data NOT in saved_file")
                pass
            pass
    def is_it_new_to_bank_data (self, sbf_dict, new_row):
        new_statement_key = new_row[statement_headings_dict["STORE"][bank_num]]

        new_amount = round(float(new_row[ statement_headings_dict["AMOUNT"][bank_num]]),2)
        pass
        new_statement_date = new_row[statement_headings_dict["DATE"][bank_num]]
        new_statement_amount = round(float (new_amount),2)
        in_xl_not_saved = self.is_statement_row_in_bd_and_not_in_saved_xl (new_statement_key)
        if new_statement_key in bank_data.keys():
            category_num_from_bank_data = bank_data[new_statement_key][sbf_headings_dict["CATEGORY"]]
            if(bank_data[new_statement_key][sbf_headings_dict["DATE"]] == new_statement_date and bank_data[new_statement_key][sbf_headings_dict["AMOUNT"]] == new_amount):
                return "BANK_DATA EXACTLY SAME AS NEW STATEMENT",new_statement_key, new_statement_amount, new_statement_date,category_num_from_bank_data
            else:
                return "EXISTS IN BANK_DATA", new_statement_key,new_statement_amount, new_statement_date, category_num_from_bank_data
        if  sbf_dict:  # has got stuff in it if true
            list_of_keys = list (sbf_dict.keys())
            my_list = "EXISTING ENTRY", list_of_keys[-1], new_statement_amount
            if new_statement_key in sbf_dict.keys():
                pass
                print ("exist")
            else:
                pass
        return "NEW ENTRY", new_statement_key, new_statement_amount, new_statement_date, "NO CATEGORY UNTIL PICKED"

    def  update_amounts(self, latest_sbf_dict, new_amount_to_check):
        pass
        new_amount_to_check = int(round(float(new_amount_to_check),2))
        print(f" ==>>> check to see amounts are different  {new_amount_to_check} vs. {latest_sbf_dict[Sbf_enum.AMOUNT.value]}<<=\n")
        if latest_sbf_dict[Sbf_enum.AMOUNT.value] != new_amount_to_check:
            latest_sbf_dict[Sbf_enum.AMOUNT.value] = new_amount_to_check
            pass

        return

class Charts:
    def __init__(self, save_file, category_dict_2):
        self.save_file = save_file
        self.category_dict_2 = category_dict_2

        cat_list = list[category_dict_2]
        pass
    # # class Open_with_openpyxl:
    # def __init(self, save_file):
    #     self.save_file = save_file
    def new_tab(self):

        wb_charts = load_workbook(save_file)
        ws_b = wb_charts.create_sheet ("Bubble Chart")
        ws_p = wb_charts.create_sheet ("Pie Chart")
        # ws_g2['A4'] = 4
        pass
        return wb_charts, ws_b, ws_p

    def pie_chart (self, wb_g2, ws_g2):
        data = [
            ['Pie', 'Sold'],
            ['Apple', 50],
            ['Cherry', 30],
            ['Pumpkin', 10],
            ['Chocolate', 40],
        ]
        wb = wb_g2
        ws = ws_g2
        # wb = Workbook()
        # ws = wb.active

        for row in data:
            ws.append(row)

        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=2, max_row=5)
        data = Reference(ws, min_col=2, min_row=1, max_row=5)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Pies sold by category"

        # Cut the first slice out of the pie
        slice = DataPoint(idx=0, explosion=20)
        pie.series[0].data_points = [slice]

        ws.add_chart(pie, "D1")

        ws = wb.create_sheet(title="Projection")

        data = [
            ['Page', 'Views'],
            ['Search', 95],
            ['Products', 4],
            ['Offers', 0.5],
            ['Sales', 0.5],
        ]

        for row in data:
            ws.append(row)

        projected_pie = ProjectedPieChart()
        projected_pie.type = "pie"
        projected_pie.splitType = "val"  # split by value
        labels = Reference(ws, min_col=1, min_row=2, max_row=5)
        data = Reference(ws, min_col=2, min_row=1, max_row=5)
        projected_pie.add_data(data, titles_from_data=True)
        projected_pie.set_categories(labels)

        ws.add_chart(projected_pie, "A10")

        from copy import deepcopy
        projected_bar = deepcopy(projected_pie)
        projected_bar.type = "bar"
        projected_bar.splitType = 'pos'  # split by position

        ws.add_chart(projected_bar, "A27")

        # wb.save("pie.xlsx")
    def bubble_chart (self, wb_g2, ws_g2):

        sheet = ws_g2
        rows = [
            ("Number of Products", "Sales in USD", "Market share"),
            (14, 12200, 15),
            (20, 60000, 33),
            (18, 24400, 10),
            (22, 32000, 42),
        ]

        for row in rows:
            sheet.append(row)

        # Create object of BubbleChart class
        chart = BubbleChart()

        # create data for plotting
        xvalues = Reference(sheet, min_col=1,
                            min_row=2, max_row=5)

        yvalues = Reference(sheet, min_col=2,
                            min_row=2, max_row=5)

        size = Reference(sheet, min_col=3,
                         min_row=2, max_row=5)

        # create a 1st series of data
        series = Series(values=yvalues, xvalues=xvalues,
                        zvalues=size, title="2013")

        # add series data to the chart object
        chart.series.append(series)
        # wb_g2.save(save_file)
        pass
        # set the title of the chart
        chart.title = " BUBBLE-CHART "

        # set the title of the x-axis
        chart.x_axis.title = " X_AXIS "

        # set the title of the y-axis
        chart.y_axis.title = " Y_AXIS "

        # add chart to the sheet
        # the top-left corner of a chart
        # is anchored to cell E2 .
        sheet.add_chart(chart, "E2")
        # wb_g2.save(save_file)



    # def new_graph_2(self):
    #     treeData = [["Type", "Leaf Color", "Height"], ["Maple", "Red", 549], ["Oak", "Green", 783],
    #                ["Pine", "Green", 1204]]
    #
    #     for row in treeData:
    #         # ws_g2.append(row)
    #         pass
    #     wb.save(save_file)
    #     pass




#**************************************************************
    # +++++++++++++++++++++++++|
    #   Start of Main          |
    # +++++++++++++++++++++++++|
#**************************************************************
bank_choice = "9"
total = 0
y= 'y'
df_g = 0
#if os.path.isfile(save_file):
#    while y == "y":
#        y = input("Graphs?")
 #       g = Graph(save_file, gws)
 #       g.new_graph()

#--------------------------------------------------
# find and load existing data file into bank_data
#--------------------------------------------------
if os.path.isfile(save_file):
    wb = load_workbook(save_file)
    ws = wb.active

    pass
    for row in list(ws.rows)[1:]:
        pass
        bank_data[row[0].value] = [c.value for c in row[1:]]
    pass
    bank_data.pop ("null", "not_found")
    ws["J2"] = "fuck"
    pass
else:  #create one
    wb = openpyxl.Workbook()
    ws = wb.active
    gws = wb.create_sheet("A graph sheet")
    sws = wb.create_sheet('category sums')
    tws = wb.create_sheet('vategory totals')
    ws.column_dimensions["B"].width = 300
    wb.save (filename=save_file)
    pass
pass
_or_f_is_it_special = False
bank_num, bank_name = intro()
month_wanted = month_intro()
bank_file_1 = Bank(bank_name_dict[bank_num], bank_num, bank_file_name_dict[bank_name], month_wanted, statement_directory)
bank_file_1.csv_to_xl()
title_data = [
    ['Bank', 'Date']
]


while bank_choice !=  "NONE" :
    pass
    total_debit = 0
    quit_or_save = False
    total_deposit = 0
    category_from_bank_data = 0
    debit = 0
    deposit = 0
    bank_file = Bank(bank_name_dict[bank_num], bank_num, bank_file_name_dict[bank_name],month_wanted, statement_directory)
    y = bank_file.get_wb_and_ws()  # tupple of file_name, wb and ws
    if y == None:
        print ("NO MONTH FOUND")
        exit(f" month: {month_wanted} not found")
    ws = y[2]   #work sheet
    wb = y[1]   #work book
    bank_file_name = y[0]

    size = Bank.get_statement_length(bank_file_name)
    starting_row_count = size[4]-1
    print (f"Pandas size at start: {size[4] + 1} ")
    key_1 = get_key_by_value(3)
#    print (f"KEY: {key_1}")
    new_amount = 0
    print (f"number of rows in statement is :{size[1]}")
#    print (type(category_sums))
#    print (category_sums)
    for tupple_new_row in ws.iter_rows(min_row=2, max_row=size[4]+1, min_col=0, max_col=9, values_only = True):
        if quit_or_save == True:
            break
        rows_this_session += 1
        bd_len = len(bank_data)
        print(f"statement rows remaining: { starting_row_count - bd_len - new_rows_completed}")
        new_row = list(tupple_new_row)
        if bank_name == "CAPITAL_ONE":
            if new_row[statement_headings_dict['AMOUNT'][bank_num]] == None:
                new_row[statement_headings_dict['AMOUNT'][bank_num]] =round(float(new_row[statement_headings_dict['PAYMENT'][bank_num]]),2)
            else:
                new_row[statement_headings_dict['AMOUNT'][bank_num]] = round(float ((new_row[statement_headings_dict['AMOUNT'][bank_num]])),2) * -1.00
        pass
        a_new_statement_row = New_statement_row( new_row, bank_name, bank_num)
 #       print (f"new_row->key is:{new_row}\n")
        exists_in_a_dictonary_or_not,the_statement_key,the_statement_amount,the_statement_date, category_from_bank_data = a_new_statement_row.is_it_new_to_bank_data(sbf_dict, new_row)
        pass
        if exists_in_a_dictonary_or_not == "BANK_DATA EXACTLY SAME AS NEW STATEMENT":
            continue
        if (exists_in_a_dictonary_or_not == "EXISTS IN BANK_DATA" ):
#           print (f"{the_statement_key} of ${the_statement_amount} Is already in bank_data with category {category_from_bank_data}")
            row_count += 1
            if the_statement_amount > 0:
                pass
                credit = the_statement_amount
            if the_statement_amount <= 0:
#                debit = abs(the_statement_amount)
                debit = (the_statement_amount)
                total_deposit += deposit
                total_debit += debit
                bank_data[the_statement_key][sbf_headings_dict["AMOUNT"]] += the_statement_amount

        if  exists_in_a_dictonary_or_not == "NEW ENTRY":
            a_new_sbf_dict_key = a_new_statement_row.new_sbf_dict()
            new_cat_inst = Add_categories()
            print (f"\nNeed category for:{list(sbf_dict.keys())[-1]} in {bank_name_dict[bank_num]} for $ {new_row[statement_headings_dict['AMOUNT'][bank_num]]}")
            new_cat_inst.print_category_menu()
            cat_num, special_cat_num = new_cat_inst.get_category()  #return the normal category number and special (btween 20 and 30

            cat_str_2 = get_key_by_value(cat_num)
            cat_num_text_str = category_dict_2[cat_str_2]
            cat_sums = Master_sum(new_amount,cat_num,sbf_dict)
            debit, deposit, category = cat_sums.fix_new_amount_signs(a_new_sbf_dict_key)
            new_cat_inst.add_new_cat_to_sbf_dict(cat_num, sbf_dict, cat_num_text_str)
            total_debit += debit
            total_deposit += deposit
#            category_sums[cat_num] += debit
            cat_str_2 = get_key_by_value(cat_num)
            category_dict_2[cat_str_2][0]  += int(debit)
            category_dict_2[cat_str_2][0]  += int(deposit)
            pass
            new_rows_completed += 1
            pass
            if cat_str_2 == '[Q]uit and Save' or  cat_str_2 == "Quit[Z]-Don't Save":
                quit_or_save = True
                bank_choice = 'NONE'
        elif (exists_in_a_dictonary_or_not == "EXISTING ENTRY" and t_or_f_is_it_special != True):
            pass
            print ("Existing Entry")
            new_amount = new_row[statement_headings_dict["AMOUNT"][bank_num]]
            a_new_statement_row.update_amounts(sbf_dict[my_list[-1]], new_amount)
            debit, deposit, category = cat_sums.fix_new_amount_signs(a_new_sbf_dict_key, total)
            print (f" Existing Vendor will add {debit} to {a_new_sbf_dict_key}'s category which is: {cat_num}")
        pass
    if quit_or_save == "Quit[Z]-Don't Save":
        print("==== Quit - not saving ====")
        sys.exit()
        pass
    s = Merge_and_save(save_file, sbf_dict, df_g)
    s.merge_data()
    # g = Graph(save_file)
    # df_g = g.new_graph()
    s.save(df_g)  # line 145
    # s.copy_chart(saved_graph_file, save_file)
    c = Charts(save_file, category_dict_2)
    wb_charts, ws_b, ws_p = c.new_tab ()
    c.bubble_chart(wb_charts, ws_b)
    c.pie_chart(wb_charts, ws_p)
    wb_charts.save(save_file)
    # g.new_graph_2()

    print ("Saved")
 #   wb.close()


  #  if special_cat_num == magic_number_quit_and_save:
  #      exit()
  #      sys.exit()
  #  print ("Reached end of Statement")
  #  print("SAVED")
  #  print("CLOSED")










